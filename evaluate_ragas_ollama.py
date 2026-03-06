#!/usr/bin/env python3
from __future__ import annotations

import argparse
import asyncio
import json
import math
import shutil
import time
from pathlib import Path
from typing import Any

import pandas as pd
from openai import AsyncOpenAI
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from ragas.llms import llm_factory
from ragas.metrics.collections.answer_accuracy import AnswerAccuracy
from ragas.metrics.collections.faithfulness.metric import Faithfulness
from ragas.metrics.collections.response_groundedness import ResponseGroundedness


DEFAULT_RESPONSE_COLUMNS = [
    "qwen_rag",
    "qwen_no_rag",
    "tinyllama_rag",
    "tinyllama_no_rag",
    "gemini_rag",
    "gemini_no_rag",
]

METRIC_NAMES = ("answer_accuracy", "faithfulness", "response_groundedness")
CACHE_VERSION = 1


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Score every model response in a comparison report with RAGAS using an Ollama judge."
        )
    )
    parser.add_argument(
        "--report",
        default="comparison_report_20260303_205035.xlsx",
        help="Path to the comparison report workbook.",
    )
    parser.add_argument(
        "--data",
        default="GNEM updated excel.xlsx",
        help="Path to the source data workbook. Used for metadata only in this script.",
    )
    parser.add_argument(
        "--judge-model",
        default="qwen2.5:14b",
        help="Ollama model used by RAGAS for scoring.",
    )
    parser.add_argument(
        "--reference-model",
        default=None,
        help="Ollama model used to generate reference answers. Defaults to --judge-model.",
    )
    parser.add_argument(
        "--base-url",
        default="http://127.0.0.1:11434/v1",
        help="OpenAI-compatible Ollama endpoint.",
    )
    parser.add_argument(
        "--api-key",
        default="ollama",
        help="API key passed to the OpenAI client. Ollama accepts any non-empty value.",
    )
    parser.add_argument(
        "--metrics",
        default="answer_accuracy,faithfulness,response_groundedness",
        help="Comma-separated subset of: answer_accuracy, faithfulness, response_groundedness",
    )
    parser.add_argument(
        "--responses",
        default=",".join(DEFAULT_RESPONSE_COLUMNS),
        help="Comma-separated response columns from the responses sheet.",
    )
    parser.add_argument(
        "--max-contexts",
        type=int,
        default=8,
        help="Maximum number of retrieved contexts to use per question.",
    )
    parser.add_argument(
        "--max-context-chars",
        type=int,
        default=4000,
        help="Maximum characters kept from each retrieved context.",
    )
    parser.add_argument(
        "--max-reference-tokens",
        type=int,
        default=1200,
        help="Max tokens for reference-answer generation.",
    )
    parser.add_argument(
        "--max-judge-tokens",
        type=int,
        default=4096,
        help="Max tokens for RAGAS judge calls.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Only evaluate the first N questions. Useful for smoke tests.",
    )
    parser.add_argument(
        "--sleep-seconds",
        type=float,
        default=0.0,
        help="Optional pause between scored responses.",
    )
    parser.add_argument(
        "--cache",
        default="ragas_eval_cache.json",
        help="Progress cache path for resumable runs.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output workbook path. Defaults to <report_stem>_ragas.xlsx",
    )
    return parser.parse_args()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    return text


def load_cache(cache_path: Path) -> dict[str, Any]:
    if not cache_path.exists():
        return {"version": CACHE_VERSION, "references": {}, "scores": {}}
    with cache_path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if data.get("version") != CACHE_VERSION:
        raise ValueError(
            f"Unsupported cache version in {cache_path}: {data.get('version')}"
        )
    data.setdefault("references", {})
    data.setdefault("scores", {})
    return data


def save_cache(cache_path: Path, data: dict[str, Any]) -> None:
    tmp_path = cache_path.with_suffix(cache_path.suffix + ".tmp")
    with tmp_path.open("w", encoding="utf-8") as handle:
        json.dump(data, handle, indent=2, ensure_ascii=True)
    tmp_path.replace(cache_path)


def build_context_map(
    retrieval_df: pd.DataFrame,
    questions: list[str],
    max_contexts: int,
    max_context_chars: int,
) -> dict[str, list[str]]:
    question_set = set(questions)
    filtered = retrieval_df[retrieval_df["question"].isin(question_set)].copy()
    filtered["rank"] = pd.to_numeric(filtered["rank"], errors="coerce").fillna(9999)
    filtered.sort_values(["question", "rank", "chunk_type"], inplace=True)

    context_map: dict[str, list[str]] = {}
    for question, group in filtered.groupby("question", sort=False):
        contexts: list[str] = []
        seen: set[str] = set()
        for _, row in group.iterrows():
            raw_text = normalize_text(row.get("text"))
            if not raw_text:
                continue
            context = (
                f"[Rank {int(row['rank'])} | {normalize_text(row.get('chunk_type'))}]\n"
                f"{raw_text[:max_context_chars]}"
            )
            if context in seen:
                continue
            seen.add(context)
            contexts.append(context)
            if len(contexts) >= max_contexts:
                break
        context_map[question] = contexts
    return context_map


def build_model_records(
    responses_df: pd.DataFrame, response_columns: list[str]
) -> tuple[list[str], list[dict[str, Any]]]:
    questions = [normalize_text(value) for value in responses_df["Question"].tolist()]
    records: list[dict[str, Any]] = []
    for _, row in responses_df.iterrows():
        question = normalize_text(row["Question"])
        if not question:
            continue
        for model in response_columns:
            records.append(
                {
                    "question": question,
                    "model": model,
                    "response": normalize_text(row.get(model)),
                }
            )
    return questions, records


def score_key(question: str, model: str) -> str:
    return f"{question}\u241f{model}"


def build_reference_prompt(question: str, contexts: list[str]) -> list[dict[str, str]]:
    context_blob = "\n\n".join(contexts) if contexts else "No retrieved context available."
    return [
        {
            "role": "system",
            "content": (
                "You write gold reference answers for evaluation. "
                "Use only the provided workbook evidence. "
                "Preserve company names, categories, locations, and numbers exactly when supported. "
                "Answer directly in plain text. "
                "If the evidence is incomplete, answer with only the supported facts. "
                "Do not invent missing values and do not explain your reasoning."
            ),
        },
        {
            "role": "user",
            "content": (
                f"Question:\n{question}\n\n"
                f"Workbook evidence:\n{context_blob}\n\n"
                "Write the best possible reference answer using only this evidence."
            ),
        },
    ]


async def generate_reference_answer(
    client: AsyncOpenAI,
    model: str,
    question: str,
    contexts: list[str],
    max_tokens: int,
) -> str:
    completion = await client.chat.completions.create(
        model=model,
        temperature=0,
        max_tokens=max_tokens,
        messages=build_reference_prompt(question, contexts),
    )
    content = completion.choices[0].message.content or ""
    return normalize_text(content)


def build_metrics(metric_names: list[str], llm: Any) -> dict[str, Any]:
    metric_map: dict[str, Any] = {}
    if "answer_accuracy" in metric_names:
        metric_map["answer_accuracy"] = AnswerAccuracy(llm=llm)
    if "faithfulness" in metric_names:
        metric_map["faithfulness"] = Faithfulness(llm=llm)
    if "response_groundedness" in metric_names:
        metric_map["response_groundedness"] = ResponseGroundedness(llm=llm)
    return metric_map


async def score_response(
    metrics: dict[str, Any],
    metric_names: list[str],
    question: str,
    response: str,
    reference: str,
    contexts: list[str],
) -> dict[str, float]:
    scores: dict[str, float] = {}
    for metric_name in metric_names:
        try:
            if metric_name == "answer_accuracy":
                result = await metrics[metric_name].ascore(
                    user_input=question,
                    response=response,
                    reference=reference,
                )
            elif metric_name == "faithfulness":
                result = await metrics[metric_name].ascore(
                    user_input=question,
                    response=response,
                    retrieved_contexts=contexts,
                )
            elif metric_name == "response_groundedness":
                result = await metrics[metric_name].ascore(
                    response=response,
                    retrieved_contexts=contexts,
                )
            else:
                raise ValueError(f"Unsupported metric: {metric_name}")
            scores[metric_name] = float(result.value)
        except Exception:
            scores[metric_name] = float("nan")
    return scores


def build_long_dataframe(
    model_records: list[dict[str, Any]],
    cache: dict[str, Any],
    metric_names: list[str],
    context_map: dict[str, list[str]],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for record in model_records:
        question = record["question"]
        model = record["model"]
        key = score_key(question, model)
        cached_scores = cache["scores"].get(key, {})
        row = {
            "question": question,
            "model": model,
            "response": record["response"],
            "reference_answer": cache["references"].get(question, ""),
            "context_count": len(context_map.get(question, [])),
        }
        for metric_name in metric_names:
            row[metric_name] = cached_scores.get(metric_name, float("nan"))
        metric_values = [row[name] for name in metric_names if pd.notna(row[name])]
        row["composite_score"] = (
            round(sum(metric_values) / len(metric_values), 6)
            if metric_values
            else float("nan")
        )
        rows.append(row)
    return pd.DataFrame(rows)


def build_wide_dataframe(long_df: pd.DataFrame, metric_names: list[str]) -> pd.DataFrame:
    wide = pd.DataFrame({"question": sorted(long_df["question"].unique())})
    for metric_name in metric_names + ["composite_score"]:
        pivot = long_df.pivot(index="question", columns="model", values=metric_name)
        pivot.columns = [f"{column}_{metric_name}" for column in pivot.columns]
        pivot = pivot.reset_index()
        wide = wide.merge(pivot, on="question", how="left")
    refs = (
        long_df[["question", "reference_answer", "context_count"]]
        .drop_duplicates(subset=["question"])
        .reset_index(drop=True)
    )
    return wide.merge(refs, on="question", how="left")


def build_summary_dataframe(long_df: pd.DataFrame, metric_names: list[str]) -> pd.DataFrame:
    summary = (
        long_df.groupby("model", dropna=False)[metric_names + ["composite_score"]]
        .mean(numeric_only=True)
        .reset_index()
    )
    counts = long_df.groupby("model", dropna=False).size().rename("questions_scored")
    summary = summary.merge(counts, on="model", how="left")
    ordered_cols = ["model", "questions_scored", *metric_names, "composite_score"]
    return summary[ordered_cols].sort_values("composite_score", ascending=False)


def build_reference_dataframe(
    questions: list[str],
    cache: dict[str, Any],
    context_map: dict[str, list[str]],
) -> pd.DataFrame:
    rows = []
    for question in questions:
        contexts = context_map.get(question, [])
        rows.append(
            {
                "question": question,
                "reference_answer": cache["references"].get(question, ""),
                "context_count": len(contexts),
                "top_context": contexts[0] if contexts else "",
            }
        )
    return pd.DataFrame(rows)


def replace_sheet_from_dataframe(workbook: Any, sheet_name: str, df: pd.DataFrame) -> None:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    worksheet = workbook.create_sheet(title=sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(row)
    for cell in worksheet[1]:
        cell.style = "Pandas"
    for column in worksheet.columns:
        max_len = 0
        letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = min(max(max_len, len(value)), 80)
        worksheet.column_dimensions[letter].width = max(14, max_len + 2)


def write_output_workbook(
    report_path: Path,
    output_path: Path,
    frames: dict[str, pd.DataFrame],
) -> None:
    shutil.copy2(report_path, output_path)
    workbook = load_workbook(output_path)
    for sheet_name, df in frames.items():
        replace_sheet_from_dataframe(workbook, sheet_name, df)
    workbook.save(output_path)


async def main() -> None:
    args = parse_args()

    report_path = Path(args.report).resolve()
    data_path = Path(args.data).resolve()
    cache_path = Path(args.cache).resolve()
    output_path = (
        Path(args.output).resolve()
        if args.output
        else report_path.with_name(f"{report_path.stem}_ragas.xlsx")
    )

    if not report_path.exists():
        raise FileNotFoundError(f"Report workbook not found: {report_path}")
    if not data_path.exists():
        raise FileNotFoundError(f"Data workbook not found: {data_path}")

    metric_names = [name.strip() for name in args.metrics.split(",") if name.strip()]
    invalid_metrics = [name for name in metric_names if name not in METRIC_NAMES]
    if invalid_metrics:
        raise ValueError(
            f"Unsupported metrics requested: {invalid_metrics}. Valid values: {METRIC_NAMES}"
        )

    response_columns = [name.strip() for name in args.responses.split(",") if name.strip()]
    if not response_columns:
        raise ValueError("At least one response column is required.")

    responses_df = pd.read_excel(report_path, sheet_name="responses")
    retrieval_df = pd.read_excel(report_path, sheet_name="retrieval")

    missing_columns = [name for name in response_columns if name not in responses_df.columns]
    if missing_columns:
        raise ValueError(f"Missing response columns in responses sheet: {missing_columns}")

    questions, model_records = build_model_records(responses_df, response_columns)
    if args.limit is not None:
        questions = questions[: args.limit]
        question_set = set(questions)
        model_records = [row for row in model_records if row["question"] in question_set]

    context_map = build_context_map(
        retrieval_df=retrieval_df,
        questions=questions,
        max_contexts=args.max_contexts,
        max_context_chars=args.max_context_chars,
    )

    cache = load_cache(cache_path)
    cache_changed = False

    async_client = AsyncOpenAI(
        base_url=args.base_url,
        api_key=args.api_key,
        timeout=300.0,
    )

    reference_model = args.reference_model or args.judge_model
    judge_llm = llm_factory(
        args.judge_model,
        client=async_client,
        temperature=0,
        max_tokens=args.max_judge_tokens,
    )
    metrics = build_metrics(metric_names, judge_llm)

    print(
        f"Questions: {len(questions)} | Responses: {len(model_records)} | "
        f"Metrics: {', '.join(metric_names)}"
    )
    print(f"Judge model: {args.judge_model}")
    print(f"Reference model: {reference_model}")
    print(f"Cache: {cache_path}")

    for index, question in enumerate(questions, start=1):
        if question in cache["references"]:
            continue
        reference_answer = await generate_reference_answer(
            client=async_client,
            model=reference_model,
            question=question,
            contexts=context_map.get(question, []),
            max_tokens=args.max_reference_tokens,
        )
        cache["references"][question] = reference_answer
        cache_changed = True
        save_cache(cache_path, cache)
        print(f"[reference {index}/{len(questions)}] generated")

    for index, record in enumerate(model_records, start=1):
        question = record["question"]
        model = record["model"]
        response = record["response"]
        key = score_key(question, model)
        cached_scores = cache["scores"].setdefault(key, {})

        if response == "":
            for metric_name in metric_names:
                cached_scores.setdefault(metric_name, float("nan"))
            cache_changed = True
            continue

        needed = [name for name in metric_names if name not in cached_scores]
        if not needed:
            continue

        scores = await score_response(
            metrics=metrics,
            metric_names=needed,
            question=question,
            response=response,
            reference=cache["references"].get(question, ""),
            contexts=context_map.get(question, []),
        )
        cached_scores.update(scores)
        cache_changed = True
        save_cache(cache_path, cache)
        print(
            f"[score {index}/{len(model_records)}] {model} | "
            f"{', '.join(f'{k}={cached_scores[k]:.3f}' if pd.notna(cached_scores[k]) else f'{k}=nan' for k in needed)}"
        )
        if args.sleep_seconds > 0:
            await asyncio.sleep(args.sleep_seconds)

    if cache_changed:
        save_cache(cache_path, cache)

    long_df = build_long_dataframe(
        model_records=model_records,
        cache=cache,
        metric_names=metric_names,
        context_map=context_map,
    )
    wide_df = build_wide_dataframe(long_df=long_df, metric_names=metric_names)
    summary_df = build_summary_dataframe(long_df=long_df, metric_names=metric_names)
    reference_df = build_reference_dataframe(
        questions=questions,
        cache=cache,
        context_map=context_map,
    )

    frames = {
        "ragas_scores_long": long_df,
        "ragas_scores_wide": wide_df,
        "ragas_summary": summary_df,
        "ragas_references": reference_df,
        "ragas_run_info": pd.DataFrame(
            [
                {
                    "created_at_unix": int(time.time()),
                    "report_path": str(report_path),
                    "data_path": str(data_path),
                    "judge_model": args.judge_model,
                    "reference_model": reference_model,
                    "metrics": ",".join(metric_names),
                    "responses_scored": len(model_records),
                    "questions_scored": len(questions),
                    "cache_path": str(cache_path),
                }
            ]
        ),
    }
    write_output_workbook(report_path=report_path, output_path=output_path, frames=frames)

    print(f"Output workbook: {output_path}")
    print(summary_df.to_string(index=False))


if __name__ == "__main__":
    asyncio.run(main())
